import openpyxl
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
import time
from bs4 import BeautifulSoup
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import os
import docx2txt
import pyautogui
from glob import glob
import re
import win32com.client as win32
from win32com.client import constants
from selenium.webdriver.support.ui import WebDriverWait


Excel_path = "C:\\Users\\akhilesh.kumar\\Desktop\\DuckCreek AK\\Forms_Finalsheet.xlsx"

wb = load_workbook(Excel_path)
ws = wb['Excess follow']


rows_count=ws.max_row
column_count=ws.max_column

policyNumber=ws.cell(row=2, column=1).value
form_numbers=[]
sheet_form_values=[]
for row_num in range(2,rows_count+1):
	x=[]
	form_numbers.append(ws.cell(row=row_num, column=3).value)
	x.append(str(ws.cell(row=row_num, column=3).value))
	x.append(str(ws.cell(row=row_num, column=7).value))
	x.append(str(ws.cell(row=row_num, column=2).value))
	sheet_form_values.append(x)
	#print(ws.cell(row=row_num, column=2).value)



caps = DesiredCapabilities.INTERNETEXPLORER
caps['ignoreProtectedModeSettings'] = True    
driver = webdriver.Ie(capabilities=caps)
##---------------For AT1-------------------

driver.get("https://ued11.duckcreekondemand.com/Policy/default.aspx")

driver.find_element_by_id("username-inputEl").send_keys("ckonapala")
driver.find_element_by_id("password-inputEl").send_keys("password")
driver.find_element_by_id("home").click()
##----------------------------------------------------------------
##--------------For UAT------------------------------------------
##driver.get("https://ueu03.duckcreekondemand.com/Policy/default.aspx")
##window_before = driver.window_handles[0]

driver.maximize_window()
time.sleep(10)

driver.find_element_by_id("quickSearchModeId-trigger-picker").click()
driver.find_element_by_xpath("//*[@id='quickSearchModeId-picker-listEl']/li[1]").click()
driver.find_element_by_id("quickSearchTextId-inputEl").send_keys(str(policyNumber))
driver.find_element_by_id("id_quickSearch").click()
time.sleep(5)
driver.find_element_by_xpath("//*[@id='quoteListLoadQuoteA']/img").click()
time.sleep(5)
driver.find_element_by_xpath("//a[contains(text(),'Pricing')]").click()
time.sleep(5)
driver.find_element_by_xpath("//span[contains(text(),'Forms')]").click()
##driver.find_element_by_name("_printJobRestrict").send_keys("abc")
time.sleep(8)
driver.find_element_by_id("perpageprintJobList-trigger-picker").click();
driver.find_element_by_xpath("//*[@id='perpageprintJobList-picker-listEl']/li[4]").click();

time.sleep(5)
soup=BeautifulSoup(driver.page_source,'lxml')
div=soup.find('div',{'id':'printJobList-body'})

tables=div.findAll('table')
table_form_values=[]
##sheet = wb['Result']
##i=2
##sheet.cell(row=1, column=2).value ="VALUE"
##sheet.cell(row=1, column=1).value ="FORM"
##for x in tables:
##    y=[]
##    #print(x.text)
##    #2==mandatory
##    #1==checked editable
##    #0 == UNCHECKED
##    f_value="NA"
##    f_name=x.tr.findAll('td')[3].text
##    if(str(x.tr.findAll('td')[0].find('input').get('value'))=="0"):
##        f_value="Unselected";
##    if(str(x.tr.findAll('td')[0].find('input').get('value'))=="1"):
##        f_value= "Selected But Editable";
##    if(str(x.tr.findAll('td')[0].find('input').get('value'))=="2"):
##        f_value= "mandatory"; 
##    y.append(str(f_name))
##    y.append(str(f_value))
##    table_form_values.append(y)
##    sheet.cell(row=i, column=1).value = str(f_name)
##    sheet.cell(row=i, column=2).value =str(f_value)
##    i=i+1
##
##
##sheet = wb['Result1']
##i=2
##sheet.cell(row=1, column=2).value ="VALUE"
##sheet.cell(row=1, column=1).value ="FORM"
##matched_form=[]
##val=[]
##for l1 in table_form_values:
##    for l2 in sheet_form_values:
##        if(l1[0]==l2[0].replace('-', '').replace('/', '')):
##            print(l1)
##            sheet.cell(row=i, column=1).value = str(l1[0])
##            sheet.cell(row=i, column=2).value =str(l1[1])
##            sheet.cell(row=i, column=3).value =str(l2[1])
##            val.append(l1[0])
##            val.append(l1[1])
##            val.append(l2[1])
##            matched_form.append(val)
##            i=i+1
##
##wb.save(path)
sheet = wb['Result']
i=2

##sheet.cell(row=1, column=1).value ="FORM"
##sheet.cell(row=1, column=2).value ="VALUE"

for x in tables:
	y=[]
	f_value="NA"
	f_name=x.findAll('td')[3].text
	if(str(x.findAll('td')[0].find('input').get('value'))=="0"):
		f_value="Unselected";
	if(str(x.findAll('td')[0].find('input').get('value'))=="1"):
		f_value= "Selected But Editable";
	if(str(x.findAll('td')[0].find('input').get('value'))=="2"):
		f_value= "mandatory";
		
	y.append(str(f_name))
	y.append(str(f_value))
	
	
	table_form_values.append(y)
##    sheet.cell(row=i, column=1).value = str(f_name)
##    sheet.cell(row=i, column=2).value =str(f_value)
##    i=i+1
##wb.save(path)

matched_form=[]

##for l1 in table_form_values:
##    for l2 in sheet_form_values:
##        if(l1[0]==l2[0].replace('-', '').replace('/', '') and l1[1]==l2[1]):
##            print(l1)
##            matched_form.append(l1)


driver.find_element_by_class_name("x-column-header-checkbox").click();
time.sleep(5)
driver.find_element_by_xpath("//*[@id='returnToPolicyA']/a/span").click()
time.sleep(5)
driver.find_element_by_link_text("Quote").click()
time.sleep(10)

## -------------------------attachment section------------------

driver.find_element_by_xpath("//span[contains(text(),'Attachments')]").click()
time.sleep(5)
#------------------------click on attachments--------------
Refresh = driver.find_element_by_xpath("//div[contains(@class,'x-tool-img')]")
Refresh.click()
time.sleep(2)
##---------------Click on Quote attachemnt-------------------------
##attachment=driver.find_element_by_xpath("(//div[div[contains(text(),'QUO - Quote Documents')]]/div/img)[1]")
def click_attachments(driver):
	attachment=driver.find_element_by_xpath("(//div[div[contains(text(),'QUO - Quote Documents')]]/div/img)[1]")
	if attachment.is_enabled:
			return attachment
	else:
			driver.find_element_by_xpath("//div[contains(@class,'x-tool-img')]").click()

TIMEOUT=30
WebDriverWait(driver, TIMEOUT, poll_frequency=0.25).until(click_attachments)
driver.find_element_by_xpath("(//div[div[contains(text(),'QUO - Quote Documents')]]/div/img)[1]").click()
time.sleep(2)
##-----------------------------------------------------
##r=1
##while r<=10:
##    print (r)
##    
##    time.sleep(1)
##    driver.find_element_by_xpath("//div[contains(@class,'x-tool-img')]").click()
##    r=r+1
##    
##        
##
##print ("---hello----")
##driver.find_element_by_xpath("(//div[div[contains(text(),'QUO - Quote Documents')]]/div/img)[1]").click()



try:
	button7location = pyautogui.locateOnScreen('save_1.png')
	while button7location== None:
		button7location = pyautogui.locateOnScreen('save_1.png')
except TypeError:
	print ("----------------exc----")

button7x, button7y = pyautogui.center(button7location)
pyautogui.click(button7x, button7y)

path= 'C:\\Users\\akhilesh.kumar\\Desktop\\DuckCreek AK\\Output\\'
files=os.listdir(path)
i=0

for file in files:
	if file.startswith('QUO_-_Quote_Documents'): 
		os.rename(os.path.join(path, file), os.path.join(path,'Quote_document.doc'))
	i = i+1

##------------------saving to doc to docx---------------------------------------##


paths = glob('C:\\Users\\akhilesh.kumar\\Desktop\\DuckCreek AK\\Output\\Quote_document.doc', recursive=True)

def save_as_docx(path):
	# Opening MS Word
	word = win32.gencache.EnsureDispatch('Word.Application')
	doc = word.Documents.Open(path)
	doc.Activate ()

	# Rename path with .docx
	new_file_abs = os.path.abspath(path)
	new_file_abs = re.sub(r'\.\w+$', '.docx', new_file_abs)

	# Save and Close
	word.ActiveDocument.SaveAs(
		new_file_abs, FileFormat=constants.wdFormatXMLDocument
	)
	doc.Close(False)
for i in paths:
	save_as_docx(i)
##--------------------------------------------------------##
test=docx2txt.process("C:\\Users\\akhilesh.kumar\\Desktop\\DuckCreek AK\\Output\\Quote_document.docx")
matched_flag=0
xcl_result=[]
tbl_result=[]
matched_forms_number=[]
#sheet = wb['Result']
for f1 in sheet_form_values:
	x_ob=[]
##    y_ob=[]
	for f2 in table_form_values:
		#if formnumber and values are same
##        if(f1[0]==f2[0] and f1[1]==f2[1]):
		if(f1[0]==f2[0]):
			x_ob.append(f1[0])
			x_ob.append(f1[1])
			x_ob.append(f2[0])
			x_ob.append(f2[1])
			if(f1[1]==f2[1]):
				x_ob.append('Yes')
			else:
				x_ob.append('No')
			x=f1[0]
			x=x[:2]+'-' +x[2:]
			x=x[:5]+'-' +x[5:]
			x=x[:9]+'-' +x[9:]
			x=x[:12]+'/' +x[12:]
			x_ob.append(test.count(x))
			matched_flag=1
			matched_forms_number.append(f1[0])
			break
	#if didn't found match in both list
	if(matched_flag==0):
		x_ob.append(f1[0])
		x_ob.append(f1[1])
		x_ob.append('NULL')
		x_ob.append('NULL')
		x_ob.append('No')
		x_ob.append(0)
		
##        y_ob.append('NULL')
##        y_ob.append('NULL')
##        y_ob.append(f22[0])
##        y_ob.append(f22[1])
##        y_ob.append('No')
##        if(f22[0]!=None or f22[0]==""):
##            x=f2[0]
##            x=x[:2]+'-' +x[2:]
##            x=x[:5]+'-' +x[5:]
##            x=x[:9]+'-' +x[9:]
##            x=x[:12]+'/'+x[12:]
##            y_ob.append(test.count(x))
##        else:
##            y_ob.append(0)
	matched_flag=0
	xcl_result.append(x_ob)
##    if  len(y_ob) != 0:
##        tbl_result.append(y_ob)
tbl_result=[]

for f1 in table_form_values:
	y_ob=[]
	if(f1[0] not in matched_forms_number):
		y_ob.append('NULL')
		y_ob.append('NULL')
		y_ob.append(f1[0])
		y_ob.append(f1[1])
		y_ob.append('No')
		if(f1[0]!=None or f1[0]==""):
			x=f1[0]
			x=x[:2]+'-' +x[2:]
			x=x[:5]+'-' +x[5:]
			x=x[:9]+'-' +x[9:]
			x=x[:12]+'/'+x[12:]
			y_ob.append(test.count(x))
		else:
			y_ob.append(0)
		tbl_result.append(y_ob)

print(xcl_result)
print(tbl_result)
sheet.cell(row=1, column=1).value = 'EXCEL FORM NUMBER-Req Sheet'
sheet.cell(row=1, column=2).value = 'EXCEL FORM CONDITION-Req Sheet'
sheet.cell(row=1, column=3).value = 'TABLE FORM NUMBER-Screen'
sheet.cell(row=1, column=4).value = 'TABLE FORM CONDITION-Screen'
sheet.cell(row=1, column=5).value = 'TABLE FORM AND EXCEL FORM MATCHED'
sheet.cell(row=1, column=6).value = 'FORM APPEARING IN PACKET'
sheet.cell(row=1, column=7).value = 'FORM NUMBER COUNT IN WORD'
i=2
for x in xcl_result:
	sheet.cell(row=i, column=1).value =str(x[0])
	sheet.cell(row=i, column=2).value =str(x[1])
	sheet.cell(row=i, column=3).value =str(x[2])
	sheet.cell(row=i, column=4).value =str(x[3])
	sheet.cell(row=i, column=5).value =str(x[4])
	sheet.cell(row=i, column=6).value =str('Yes' if int(x[5])>=2 else 'No')
	sheet.cell(row=i, column=7).value =str(x[5])
	i=i+1
	
for xy in tbl_result:
	sheet.cell(row=i, column=1).value =str(xy[0])
	sheet.cell(row=i, column=2).value =str(xy[1])
	sheet.cell(row=i, column=3).value =str(xy[2])
	sheet.cell(row=i, column=4).value =str(xy[3])
	sheet.cell(row=i, column=5).value =str(xy[4])
	sheet.cell(row=i, column=6).value =str('Yes' if int(xy[5])>=2 else 'No')
	sheet.cell(row=i, column=7).value =str(xy[5])
	i=i+1
	
wb.save(Excel_path)

print ("-------------Completed---------------")
